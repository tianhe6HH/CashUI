"""月度财务报表生成（正规财务报表格式，按科目专款专用）。"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select, func
from sqlalchemy.orm import joinedload

from app.config import settings
from app.database import SessionLocal
from app.models import Account, Transaction, TransactionType, Report


def _month_range(month: str):
    """根据 YYYY-MM 返回起止日期字符串（用于 SQLite strftime 过滤）。"""
    start = f"{month}-01"
    y, m = int(month[:4]), int(month[5:7])
    if m == 12:
        end = f"{y + 1}-01-01"
    else:
        end = f"{y}-{m + 1:02d}-01"
    return start, end


def _sum(db, account_id, ttype, start, end):
    """指定科目、类型、时间段的金额合计。"""
    conds = [Transaction.account_id == account_id, Transaction.type == ttype]
    if start is not None:
        conds.append(Transaction.created_at >= start)
    if end is not None:
        conds.append(Transaction.created_at < end)
    return float(
        db.scalar(select(func.coalesce(func.sum(Transaction.amount), 0)).where(*conds))
        or 0
    )


def generate_monthly_report(month: str) -> Path:
    """生成指定月份的财务报表，返回文件路径。"""
    db = SessionLocal()
    try:
        start, end = _month_range(month)
        accounts = db.scalars(select(Account).order_by(Account.id)).all()

        wb = Workbook()

        # ============ Sheet1：收支汇总（按科目） ============
        ws = wb.active
        ws.title = "收支汇总"
        headers = ["科目", "期初结余", "本期收入", "本期支出", "期末结余"]
        ws.append(headers)

        total_open = total_in = total_out = total_close = 0.0
        for acc in accounts:
            hist_in = _sum(db, acc.id, TransactionType.income, None, start)
            hist_out = _sum(db, acc.id, TransactionType.expense, None, start)
            opening = round(hist_in - hist_out, 2)
            month_in = round(_sum(db, acc.id, TransactionType.income, start, end), 2)
            month_out = round(_sum(db, acc.id, TransactionType.expense, start, end), 2)
            closing = round(opening + month_in - month_out, 2)
            ws.append([acc.name, opening, month_in, month_out, closing])
            total_open += opening
            total_in += month_in
            total_out += month_out
            total_close += closing

        # 合计行
        ws.append(
            ["合计", round(total_open, 2), round(total_in, 2),
             round(total_out, 2), round(total_close, 2)]
        )

        # ============ Sheet2：收入明细 ============
        ws_in = wb.create_sheet("收入明细")
        ws_in.append(["日期", "科目", "缴款人", "金额", "备注"])
        rows_in = db.scalars(
            select(Transaction)
            .options(joinedload(Transaction.account), joinedload(Transaction.funder))
            .where(
                Transaction.type == TransactionType.income,
                Transaction.created_at >= start,
                Transaction.created_at < end,
            )
            .order_by(Transaction.created_at, Transaction.id)
        ).all()
        for t in rows_in:
            ws_in.append(
                [
                    t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
                    t.account.name if t.account else "",
                    t.funder.name if t.funder else "",
                    float(t.amount),
                    t.note,
                ]
            )

        # ============ Sheet3：支出明细 ============
        ws_out = wb.create_sheet("支出明细")
        ws_out.append(["日期", "科目", "金额", "备注"])
        rows_out = db.scalars(
            select(Transaction)
            .options(joinedload(Transaction.account))
            .where(
                Transaction.type == TransactionType.expense,
                Transaction.created_at >= start,
                Transaction.created_at < end,
            )
            .order_by(Transaction.created_at, Transaction.id)
        ).all()
        for t in rows_out:
            ws_out.append(
                [
                    t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
                    t.account.name if t.account else "",
                    float(t.amount),
                    t.note,
                ]
            )

        # ============ 样式（表头加粗 + 合计行高亮） ============
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        for sheet in (ws, ws_in, ws_out):
            for c in sheet[1]:
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center")
        # 汇总表合计行高亮
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)

        # ============ 保存 ============
        out_dir = Path(settings.REPORT_DIR)
        out_dir.mkdir(parents=True, exist_ok=True)
        file_path = out_dir / f"{month}.xlsx"
        wb.save(file_path)

        # 记录到数据库
        existing = db.scalar(select(Report).where(Report.month == month))
        if existing is None:
            existing = Report(month=month, file_path=str(file_path))
            db.add(existing)
        else:
            existing.file_path = str(file_path)
            existing.generated_at = datetime.now()
        db.commit()
        return file_path
    finally:
        db.close()
